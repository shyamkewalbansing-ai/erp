"""
Excel Export Module - Export van boekhouding data naar Excel
"""
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_styled_workbook() -> Workbook:
    """Maak een nieuwe workbook met standaard styling"""
    return Workbook()


def style_header_row(ws, row: int = 1, num_cols: int = 10):
    """Stijl de header rij"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def auto_column_width(ws):
    """Pas kolombreedte automatisch aan"""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


def export_grootboek_excel(rekeningen: List[Dict[str, Any]], journaalposten: List[Dict[str, Any]]) -> bytes:
    """Export grootboek naar Excel"""
    wb = create_styled_workbook()
    
    # Rekeningen sheet
    ws1 = wb.active
    ws1.title = "Rekeningschema"
    
    headers = ["Code", "Naam", "Type", "Categorie", "Valuta", "Saldo"]
    for col, header in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=header)
    style_header_row(ws1, 1, len(headers))
    
    for row, rek in enumerate(rekeningen, 2):
        ws1.cell(row=row, column=1, value=rek.get('code', ''))
        ws1.cell(row=row, column=2, value=rek.get('naam', ''))
        ws1.cell(row=row, column=3, value=rek.get('type', ''))
        ws1.cell(row=row, column=4, value=rek.get('categorie', ''))
        ws1.cell(row=row, column=5, value=rek.get('valuta', 'SRD'))
        ws1.cell(row=row, column=6, value=rek.get('saldo', 0))
    
    auto_column_width(ws1)
    
    # Journaalposten sheet
    ws2 = wb.create_sheet("Journaalposten")
    headers2 = ["Volgnummer", "Datum", "Dagboek", "Omschrijving", "Debet", "Credit", "Status"]
    for col, header in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=header)
    style_header_row(ws2, 1, len(headers2))
    
    for row, post in enumerate(journaalposten, 2):
        ws2.cell(row=row, column=1, value=post.get('volgnummer', ''))
        ws2.cell(row=row, column=2, value=post.get('datum', ''))
        ws2.cell(row=row, column=3, value=post.get('dagboek_code', ''))
        ws2.cell(row=row, column=4, value=post.get('omschrijving', ''))
        ws2.cell(row=row, column=5, value=post.get('totaal_debet', 0))
        ws2.cell(row=row, column=6, value=post.get('totaal_credit', 0))
        ws2.cell(row=row, column=7, value=post.get('status', ''))
    
    auto_column_width(ws2)
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_debiteuren_excel(debiteuren: List[Dict[str, Any]], facturen: List[Dict[str, Any]]) -> bytes:
    """Export debiteuren en verkoopfacturen naar Excel"""
    wb = create_styled_workbook()
    
    # Debiteuren sheet
    ws1 = wb.active
    ws1.title = "Debiteuren"
    
    headers = ["Nummer", "Naam", "Adres", "Plaats", "Telefoon", "Email", "Valuta", "Openstaand"]
    for col, header in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=header)
    style_header_row(ws1, 1, len(headers))
    
    for row, deb in enumerate(debiteuren, 2):
        ws1.cell(row=row, column=1, value=deb.get('nummer', ''))
        ws1.cell(row=row, column=2, value=deb.get('naam', ''))
        ws1.cell(row=row, column=3, value=deb.get('adres', ''))
        ws1.cell(row=row, column=4, value=deb.get('plaats', ''))
        ws1.cell(row=row, column=5, value=deb.get('telefoon', ''))
        ws1.cell(row=row, column=6, value=deb.get('email', ''))
        ws1.cell(row=row, column=7, value=deb.get('valuta', 'SRD'))
        ws1.cell(row=row, column=8, value=deb.get('openstaand_bedrag', 0))
    
    auto_column_width(ws1)
    
    # Facturen sheet
    ws2 = wb.create_sheet("Verkoopfacturen")
    headers2 = ["Factuurnummer", "Datum", "Klant", "Vervaldatum", "Subtotaal", "BTW", "Totaal", "Openstaand", "Status"]
    for col, header in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=header)
    style_header_row(ws2, 1, len(headers2))
    
    for row, fact in enumerate(facturen, 2):
        ws2.cell(row=row, column=1, value=fact.get('factuurnummer', ''))
        ws2.cell(row=row, column=2, value=fact.get('factuurdatum', ''))
        ws2.cell(row=row, column=3, value=fact.get('debiteur_naam', ''))
        ws2.cell(row=row, column=4, value=fact.get('vervaldatum', ''))
        ws2.cell(row=row, column=5, value=fact.get('subtotaal', 0))
        ws2.cell(row=row, column=6, value=fact.get('btw_bedrag', 0))
        ws2.cell(row=row, column=7, value=fact.get('totaal_incl_btw', 0))
        ws2.cell(row=row, column=8, value=fact.get('openstaand_bedrag', 0))
        ws2.cell(row=row, column=9, value=fact.get('status', ''))
    
    auto_column_width(ws2)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_crediteuren_excel(crediteuren: List[Dict[str, Any]], facturen: List[Dict[str, Any]]) -> bytes:
    """Export crediteuren en inkoopfacturen naar Excel"""
    wb = create_styled_workbook()
    
    # Crediteuren sheet
    ws1 = wb.active
    ws1.title = "Crediteuren"
    
    headers = ["Nummer", "Naam", "Adres", "Plaats", "Telefoon", "Email", "Valuta", "Openstaand"]
    for col, header in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=header)
    style_header_row(ws1, 1, len(headers))
    
    for row, cred in enumerate(crediteuren, 2):
        ws1.cell(row=row, column=1, value=cred.get('nummer', ''))
        ws1.cell(row=row, column=2, value=cred.get('naam', ''))
        ws1.cell(row=row, column=3, value=cred.get('adres', ''))
        ws1.cell(row=row, column=4, value=cred.get('plaats', ''))
        ws1.cell(row=row, column=5, value=cred.get('telefoon', ''))
        ws1.cell(row=row, column=6, value=cred.get('email', ''))
        ws1.cell(row=row, column=7, value=cred.get('valuta', 'SRD'))
        ws1.cell(row=row, column=8, value=cred.get('openstaand_bedrag', 0))
    
    auto_column_width(ws1)
    
    # Facturen sheet
    ws2 = wb.create_sheet("Inkoopfacturen")
    headers2 = ["Intern Nr", "Extern Nr", "Datum", "Leverancier", "Vervaldatum", "Subtotaal", "BTW", "Totaal", "Openstaand", "Status"]
    for col, header in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=header)
    style_header_row(ws2, 1, len(headers2))
    
    for row, fact in enumerate(facturen, 2):
        ws2.cell(row=row, column=1, value=fact.get('intern_nummer', ''))
        ws2.cell(row=row, column=2, value=fact.get('extern_factuurnummer', ''))
        ws2.cell(row=row, column=3, value=fact.get('factuurdatum', ''))
        ws2.cell(row=row, column=4, value=fact.get('crediteur_naam', ''))
        ws2.cell(row=row, column=5, value=fact.get('vervaldatum', ''))
        ws2.cell(row=row, column=6, value=fact.get('subtotaal', 0))
        ws2.cell(row=row, column=7, value=fact.get('btw_bedrag', 0))
        ws2.cell(row=row, column=8, value=fact.get('totaal_incl_btw', 0))
        ws2.cell(row=row, column=9, value=fact.get('openstaand_bedrag', 0))
        ws2.cell(row=row, column=10, value=fact.get('status', ''))
    
    auto_column_width(ws2)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_btw_aangifte_excel(rapport: Dict[str, Any]) -> bytes:
    """Export BTW aangifte naar Excel"""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "BTW Aangifte"
    
    # Titel
    ws.cell(row=1, column=1, value="BTW AANGIFTE SURINAME")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws.cell(row=2, column=1, value=f"Periode: {rapport.get('periode', {}).get('jaar', '')}, Kwartaal {rapport.get('periode', {}).get('kwartaal', '')}")
    ws.cell(row=3, column=1, value=f"Datum: {datetime.now().strftime('%d-%m-%Y')}")
    
    # Headers
    ws.cell(row=5, column=1, value="Omschrijving")
    ws.cell(row=5, column=2, value="Bedrag (SRD)")
    style_header_row(ws, 5, 2)
    
    # Data
    data = [
        ("BTW over verkopen (af te dragen)", rapport.get('btw_verkoop', 0)),
        ("BTW over inkopen (voorbelasting)", rapport.get('btw_inkoop', 0)),
        ("", ""),
        ("Te betalen / Te vorderen", rapport.get('btw_te_betalen', 0))
    ]
    
    for row, (omschr, bedrag) in enumerate(data, 6):
        ws.cell(row=row, column=1, value=omschr)
        ws.cell(row=row, column=2, value=bedrag)
        if omschr == "Te betalen / Te vorderen":
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)
    
    auto_column_width(ws)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_winst_verlies_excel(rapport: Dict[str, Any]) -> bytes:
    """Export Winst & Verlies rekening naar Excel"""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Winst en Verlies"
    
    # Titel
    ws.cell(row=1, column=1, value="WINST & VERLIES REKENING")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')
    
    ws.cell(row=2, column=1, value=f"Jaar: {rapport.get('jaar', datetime.now().year)}")
    
    # Omzet sectie
    ws.cell(row=4, column=1, value="OMZET")
    ws.cell(row=4, column=1).font = Font(bold=True)
    style_header_row(ws, 4, 2)
    
    row = 5
    for item in rapport.get('omzet', []):
        ws.cell(row=row, column=1, value=item.get('naam', ''))
        ws.cell(row=row, column=2, value=item.get('bedrag', 0))
        row += 1
    
    ws.cell(row=row, column=1, value="Totaal Omzet")
    ws.cell(row=row, column=2, value=rapport.get('totaal_omzet', 0))
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 2
    
    # Kosten sectie
    ws.cell(row=row, column=1, value="KOSTEN")
    ws.cell(row=row, column=1).font = Font(bold=True)
    style_header_row(ws, row, 2)
    row += 1
    
    for item in rapport.get('kosten', []):
        ws.cell(row=row, column=1, value=item.get('naam', ''))
        ws.cell(row=row, column=2, value=item.get('bedrag', 0))
        row += 1
    
    ws.cell(row=row, column=1, value="Totaal Kosten")
    ws.cell(row=row, column=2, value=rapport.get('totaal_kosten', 0))
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 2
    
    # Resultaat
    ws.cell(row=row, column=1, value="NETTO WINST")
    ws.cell(row=row, column=2, value=rapport.get('netto_winst', 0))
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=2).font = Font(bold=True, size=12)
    
    auto_column_width(ws)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_balans_excel(rapport: Dict[str, Any]) -> bytes:
    """Export Balans naar Excel"""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Balans"
    
    # Titel
    ws.cell(row=1, column=1, value="BALANS")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws.cell(row=2, column=1, value=f"Datum: {rapport.get('datum', datetime.now().strftime('%d-%m-%Y'))}")
    
    # Activa
    ws.cell(row=4, column=1, value="ACTIVA")
    ws.cell(row=4, column=1).font = Font(bold=True)
    ws.cell(row=4, column=3, value="PASSIVA")
    ws.cell(row=4, column=3).font = Font(bold=True)
    
    # Headers
    ws.cell(row=5, column=1, value="Code")
    ws.cell(row=5, column=2, value="Saldo")
    ws.cell(row=5, column=3, value="Code")
    ws.cell(row=5, column=4, value="Saldo")
    style_header_row(ws, 5, 4)
    
    activa = rapport.get('activa', [])
    passiva = rapport.get('passiva', [])
    max_rows = max(len(activa), len(passiva))
    
    for i in range(max_rows):
        row = 6 + i
        if i < len(activa):
            ws.cell(row=row, column=1, value=activa[i].get('naam', ''))
            ws.cell(row=row, column=2, value=activa[i].get('saldo', 0))
        if i < len(passiva):
            ws.cell(row=row, column=3, value=passiva[i].get('naam', ''))
            ws.cell(row=row, column=4, value=passiva[i].get('saldo', 0))
    
    # Totalen
    total_row = 6 + max_rows + 1
    ws.cell(row=total_row, column=1, value="TOTAAL ACTIVA")
    ws.cell(row=total_row, column=2, value=rapport.get('totaal_activa', 0))
    ws.cell(row=total_row, column=3, value="TOTAAL PASSIVA")
    ws.cell(row=total_row, column=4, value=rapport.get('totaal_passiva', 0))
    
    for col in range(1, 5):
        ws.cell(row=total_row, column=col).font = Font(bold=True)
    
    auto_column_width(ws)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_ouderdom_excel(rapport: Dict[str, Any], type_label: str = "Debiteuren") -> bytes:
    """Export Ouderdomsanalyse naar Excel"""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = f"Ouderdom {type_label}"
    
    # Titel
    ws.cell(row=1, column=1, value=f"OUDERDOMSANALYSE {type_label.upper()}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    ws.cell(row=2, column=1, value=f"Datum: {datetime.now().strftime('%d-%m-%Y')}")
    
    # Headers
    headers = ["0-30 dagen", "31-60 dagen", "61-90 dagen", ">90 dagen", "Totaal"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(headers))
    
    # Data
    ws.cell(row=5, column=1, value=rapport.get('0_30', 0))
    ws.cell(row=5, column=2, value=rapport.get('31_60', 0))
    ws.cell(row=5, column=3, value=rapport.get('61_90', 0))
    ws.cell(row=5, column=4, value=rapport.get('90_plus', 0))
    ws.cell(row=5, column=5, value=rapport.get('totaal', 0))
    
    ws.cell(row=5, column=5).font = Font(bold=True)
    
    auto_column_width(ws)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
